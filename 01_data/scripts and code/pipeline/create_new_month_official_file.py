"""
create_new_month_official_file.py
====================================
เพิ่ม 2026-08-02 — สร้างไฟล์ "บัญชีน้ำ" ทางการของเดือนใหม่ (01_data/Reservoirs/inflow/<year>/
<year>_<month>_MNR.xlsx) จากไฟล์เดือนก่อนหน้า เพื่อแก้ manual dependency ที่บันทึกไว้ใน
ARCHITECTURE.md > "Manual Dependencies" (ยืนยันเกิดขึ้นจริงแล้ว 2 ครั้ง: มิ.ย.->ก.ค. เมื่อ 2026-07-05
และ ก.ค.->ส.ค. เมื่อ 2026-08-02 -- ทั้ง reservoir_official_file_writer.py (เขียนไฟล์ทางการรายวัน)
และ data_pipeline.py::_ri_load_raw_monthly_data() (โมเดลพยากรณ์บน Colab) ต้องมีไฟล์เดือนนั้นๆ อยู่ก่อน
ถึงจะทำงานกับเดือนนั้นได้ -- ไม่มี auto-create ไฟล์ใหม่ในระบบเดิม)

=== การทำงาน ===

  1. คัดลอกไฟล์เดือนก่อนหน้า (ทั้งไฟล์ ด้วย shutil.copy2 -- คง formatting/สไตล์/merged cells ทุกอย่าง
     ไว้เป๊ะ เพราะไฟล์นี้ไม่มี formula เหลืออยู่แล้ว -- reservoir_official_file_writer.py flatten
     formula เป็นค่าคงที่ทุกครั้งที่เขียนแถวใหม่ ยืนยันจากการตรวจสอบไฟล์จริง 2026-08-02 ว่าไม่มี cell
     ไหนเป็น "=..." เหลืออยู่เลยทั้งชีต)
  2. แก้ cell หัวเดือน (E3 ตาม layout ที่ยืนยันแล้ว: D3="เดือน", E3=<ชื่อเดือนไทย>, F3="ปี", G3=<ปี พ.ศ.>)
     -- ปีพ.ศ. (G3) ไม่เปลี่ยนถ้ายังอยู่ปีเดียวกัน (เปลี่ยนเฉพาะกรณีข้ามปีใหม่ -- สคริปต์นี้คำนวณให้เอง
     จาก target_year)
  3. ล้างข้อมูล 10 คอลัมน์ (B ถึง K: level/storage/inflow/O/spill/rain_mm/rain_m3/evap/infiltration/
     delta_s) ในทุกแถวข้อมูล (FIRST_DATA_ROW=6 ถึงแถวสุดท้ายของเดือนใหม่ตามจำนวนวันจริง) -- เหลือแค่
     คอลัมน์ A (เลขวันที่ 1-N) ไว้ให้ reservoir_official_file_writer.py เข้ามาเติมทีละวันตามปกติ
  4. ไม่แตะแถว footer (แถว 40 ในไฟล์ต้นแบบ -- ข้อความอ้างอิงแหล่งข้อมูลทั่วไป ไม่ผูกกับเดือนใดเดือนหนึ่ง)

**ไม่แตะไฟล์ต้นฉบับ (เดือนก่อนหน้า) เลย** -- อ่านอย่างเดียว, เขียนแค่ไฟล์เดือนใหม่ที่ยังไม่มีอยู่ก่อน
(raise FileExistsError ถ้าไฟล์ปลายทางมีอยู่แล้ว กันเขียนทับของจริงโดยไม่ตั้งใจ)

=== วิธีใช้ ===

    python create_new_month_official_file.py --year 2026 --month 8 --from-year 2026 --from-month 7

    (--from-year/--from-month ไม่ระบุ = ใช้เดือนก่อนหน้าของ --year/--month โดยอัตโนมัติ)

=== หลังรันแล้วต้องทำอะไรต่อ ===

  - ไฟล์ใหม่พร้อมให้ reservoir_official_file_writer.py / reservoir_daily_orchestration.py เขียนข้อมูล
    รายวันของเดือนใหม่ได้ทันทีในรอบ schtasks ถัดไป (07:30) -- ไม่ต้อง restart/ตั้งค่าอะไรเพิ่ม
  - Colab (data_pipeline_colab.py) จะเห็นไฟล์ใหม่นี้หลัง sync_to_drive.bat รอบถัดไป sync ขึ้น Drive
    (ไฟล์เปล่า/มีแค่เลขวันที่ยังไม่กระทบ -- _ri_load_raw_monthly_data() จะข้ามแถวที่ยังไม่มีข้อมูลไปเอง
    เหมือนที่จัดการแถวว่างท้ายเดือนของไฟล์เดิมอยู่แล้ว)
"""

from __future__ import annotations

import argparse
import calendar
import datetime as dt
import shutil
import sys
from pathlib import Path

import openpyxl

RESERVOIR_INFLOW_RAW_DIR = Path(__file__).resolve().parent.parent.parent / "Reservoirs" / "inflow"

MONTH_NUM_TO_ENGLISH_NAME = {
    1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June",
    7: "July", 8: "August", 9: "September", 10: "October", 11: "November", 12: "December",
}
MONTH_NUM_TO_THAI_NAME = {
    1: "มกราคม", 2: "กุมภาพันธ์", 3: "มีนาคม", 4: "เมษายน", 5: "พฤษภาคม", 6: "มิถุนายน",
    7: "กรกฎาคม", 8: "สิงหาคม", 9: "กันยายน", 10: "ตุลาคม", 11: "พฤศจิกายน", 12: "ธันวาคม",
}

SHEET_NAME = "บัญชีน้ำ"
FIRST_DATA_ROW = 6  # แถวของวันที่ 1 -- ตรงกับ reservoir_official_file_writer.py
DATA_COLS = list(range(2, 12))  # B..K (level, storage, inflow, O, spill, rain_mm, rain_m3, evap, infiltration, delta_s)
MONTH_LABEL_CELL = "E3"
YEAR_LABEL_CELL = "G3"


def official_file_path(year: int, month: int) -> Path:
    month_name = MONTH_NUM_TO_ENGLISH_NAME[month]
    return RESERVOIR_INFLOW_RAW_DIR / str(year) / f"{year}_{month_name}_MNR.xlsx"


def _prev_month(year: int, month: int) -> tuple[int, int]:
    if month == 1:
        return year - 1, 12
    return year, month - 1


def create_new_month_file(
    target_year: int, target_month: int,
    from_year: int | None = None, from_month: int | None = None,
) -> Path:
    if from_year is None or from_month is None:
        from_year, from_month = _prev_month(target_year, target_month)

    src_path = official_file_path(from_year, from_month)
    dst_path = official_file_path(target_year, target_month)

    if not src_path.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ต้นแบบ: {src_path}")
    if dst_path.exists():
        raise FileExistsError(f"ไฟล์ปลายทางมีอยู่แล้ว (ไม่เขียนทับ): {dst_path}")

    dst_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src_path, dst_path)
    print(f"[INFO] คัดลอก {src_path} -> {dst_path}")

    wb = openpyxl.load_workbook(dst_path, data_only=False)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"ไม่พบชีต '{SHEET_NAME}' ใน {dst_path}")
    ws = wb[SHEET_NAME]

    # เดือน/ปี พ.ศ. -- G3 เปลี่ยนเฉพาะข้ามปีใหม่ (target_year != from_year)
    ws[MONTH_LABEL_CELL] = MONTH_NUM_TO_THAI_NAME[target_month]
    thai_year = target_year + 543
    ws[YEAR_LABEL_CELL] = thai_year
    print(f"[INFO] ตั้งหัวเดือน: {MONTH_LABEL_CELL}={MONTH_NUM_TO_THAI_NAME[target_month]!r} {YEAR_LABEL_CELL}={thai_year}")

    n_days = calendar.monthrange(target_year, target_month)[1]
    last_row = FIRST_DATA_ROW + n_days - 1

    for row in range(FIRST_DATA_ROW, last_row + 1):
        for col in DATA_COLS:
            ws.cell(row=row, column=col).value = None
    print(f"[INFO] ล้างข้อมูลคอลัมน์ B-K แถว {FIRST_DATA_ROW}-{last_row} ({n_days} วัน) -- เหลือแค่เลขวันที่ในคอลัมน์ A")

    # ถ้าเดือนต้นแบบมีวันมากกว่าเดือนใหม่ (เช่น ก.ค. 31 วัน -> เดือนใหม่ 30 วัน) ต้องล้างแถวส่วนเกินทิ้ง
    # ทั้งแถว (รวมคอลัมน์ A ด้วย) ไม่ให้มีเลขวันที่ผีที่ไม่มีจริงค้างอยู่
    src_wb_check = openpyxl.load_workbook(src_path, data_only=False)
    src_ws_check = src_wb_check[SHEET_NAME]
    row_probe = last_row + 1
    while src_ws_check.cell(row=row_probe, column=1).value is not None and row_probe < last_row + 5:
        for col in range(1, 12):
            ws.cell(row=row_probe, column=col).value = None
        print(f"[INFO] ล้างแถวส่วนเกิน {row_probe} (เดือนใหม่มีวันน้อยกว่าเดือนต้นแบบ)")
        row_probe += 1

    wb.save(dst_path)
    print(f"[OK] บันทึก {dst_path} เรียบร้อย ({n_days} วัน พร้อมให้ reservoir_official_file_writer.py เขียนข้อมูลรายวัน)")
    return dst_path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--year", type=int, required=True)
    parser.add_argument("--month", type=int, required=True)
    parser.add_argument("--from-year", type=int, default=None)
    parser.add_argument("--from-month", type=int, default=None)
    args = parser.parse_args(argv)

    create_new_month_file(args.year, args.month, args.from_year, args.from_month)
    return 0


if __name__ == "__main__":
    sys.exit(main())
