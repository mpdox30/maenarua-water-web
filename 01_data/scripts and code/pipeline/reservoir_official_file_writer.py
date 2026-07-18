"""
reservoir_official_file_writer.py
===================================
2026-07-18 เพิ่ม — เขียนผลคำนวณจาก reservoir_daily_orchestration.py ลงไฟล์ "บัญชีน้ำ" ทางการจริง
(01_data/Reservoirs/inflow/<year>/<year>_<month>_MNR.xlsx) แทนที่จะเขียนแค่ shadow-mode CSV แยก
ต่างหากเหมือนเดิม — ผู้ใช้ยืนยันแล้ว (2026-07-18) ให้ไป live ทันที โดยยอมรับความเสี่ยงคลาดเคลื่อน
บางวันจาก rain-window convention (07:00→07:00) ที่ยังไม่ยืนยัน 100%

=== ⚠️ ข้อจำกัดสำคัญที่ต้องรู้ก่อนแก้ไฟล์นี้ ===

**openpyxl เขียนทับไฟล์ .xlsx แบบมี formula ไม่ได้โดยไม่ทำลาย cached value ของ formula เดิมทั้งไฟล์**
ทดสอบแล้ว (2026-07-18): เปิดไฟล์ 2026_July_MNR.xlsx ด้วย openpyxl (data_only=False) แล้ว save
กลับ โดยไม่แก้อะไรเลย — cached value ของ **ทุก formula cell ในทุกชีต** (ยืนยัน 10,726 cell)
หายหมด (data_only=True อ่านได้ None ทั้งหมด) เพราะ openpyxl ไม่ได้ evaluate formula เอง แค่เก็บ
สตริงสูตรไว้ ส่วน cached value ที่ Excel เคยคำนวณไว้ให้จะหายไปตอน save ถ้าไม่ได้เปิดผ่าน Excel/
LibreOffice จริงมา recalculate — ซึ่ง pandas.read_excel() (ที่ _ri_load_raw_monthly_data() ใน
data_pipeline.py ใช้) อ่านแบบ data_only เสมอ ถ้าไม่แก้ปัญหานี้ก่อน จะทำให้ pipeline หลักอ่านข้อมูล
ย้อนหลังทั้งเดือนไม่ได้เลยหลัง save ครั้งแรก (ไม่ใช่แค่แถวที่เพิ่งเขียนใหม่)

**วิธีแก้ที่ใช้ในไฟล์นี้**: ก่อนเขียนแถวใหม่ทุกครั้ง จะ "flatten" สูตรทั้งหมดในไฟล์เป็นค่าตัวเลข
ล้วนก่อน (อ่าน cached value ของทุก cell ที่เป็นสูตรด้วย data_only=True แล้วเขียนทับด้วยค่านั้นตรงๆ
แทนสตริงสูตร) — ทำให้ไฟล์ไม่พึ่ง Excel recalculation อีกต่อไป (ตัวเลขถูกต้องเหมือนเดิมทุกจุด แค่ไม่ใช่
"live formula" ที่แก้ B แล้ว C/D/H/I/J/K จะขยับตามอัตโนมัติถ้าเปิดใน Excel — เป็น trade-off ที่จำเป็น
สำหรับให้ pipeline อัตโนมัติเขียนไฟล์นี้ได้อย่างปลอดภัย) ทดสอบแล้วว่า **ทุก formula cell ในไฟล์มี
cached value ให้ fallback ครบ 100%** (ไม่มี cell ไหนเป็นสูตรที่ไม่เคยถูกคำนวณมาก่อนเลย)

ทุกครั้งที่เรียก write_computed_days() จะ backup ไฟล์เดิมไว้ก่อนเสมอ (path .bak_before_live_write_<timestamp>)

=== Layout ของชีต "บัญชีน้ำ" (ยืนยันจาก 2026_July_MNR.xlsx) ===

แถว 6 = วันที่ 1 ของเดือน, แถว N+5 = วันที่ N (คอลัมน์ A มีเลขวันที่ล่วงหน้าเต็มทุกแถวถึงวันสุดท้าย
ของเดือนอยู่แล้ว แม้แถวที่ยังไม่มีข้อมูลจริง — เช็คว่า "มีข้อมูลจริง" ให้ดูคอลัมน์ B ไม่ใช่ A)

| คอลัมน์ | ชื่อ | ที่มา (เขียนจากไฟล์นี้) |
|---|---|---|
| A | Date (เลขวันที่) | day number |
| B | Water Level (MSL) | compute_for_date()["level_msl"] |
| C | Water Volume (M3) | compute_for_date()["storage_m3"] |
| D | Inflow (M3) | compute_for_date()["inflow_m3"] |
| E | O (ปล่อยออก) | compute_for_date()["release_o_m3"] |
| F | Spill | compute_for_date()["spill_m3"] |
| G | Rain 24h (mm) | compute_for_date()["rain_24h_mm"] |
| H | R runoff (M3) | คำนวณซ้ำจาก reservoir_water_balance.compute_daily_row() (ไม่ได้อยู่ใน compute_for_date() ตรงๆ) |
| I | Evap (M3) | compute_for_date()["evap_m3"] |
| J | Infiltration (M3) | compute_for_date()["infiltration_m3"] |
| K | ΔS (M3) | compute_for_date()["delta_s_m3"] |
"""

from __future__ import annotations

import datetime as dt
import logging
import shutil
import sys
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parent))

import reservoir_daily_orchestration as rdo
import reservoir_water_balance as rwb

logger = logging.getLogger("reservoir_official_file_writer")

RESERVOIR_INFLOW_RAW_DIR = Path(__file__).resolve().parent.parent.parent / "Reservoirs" / "inflow"
SHEET_NAME = "บัญชีน้ำ"
FIRST_DATA_ROW = 6  # แถวของวันที่ 1

MONTH_NUM_TO_ENGLISH_NAME = {
    1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June",
    7: "July", 8: "August", 9: "September", 10: "October", 11: "November", 12: "December",
}

COL = {"day": 1, "level": 2, "storage": 3, "inflow": 4, "O": 5, "spill": 6,
       "rain_mm": 7, "rain_m3": 8, "evap": 9, "infiltration": 10, "delta_s": 11}


def official_file_path(target_date: dt.date) -> Path:
    month_name = MONTH_NUM_TO_ENGLISH_NAME[target_date.month]
    return RESERVOIR_INFLOW_RAW_DIR / str(target_date.year) / f"{target_date.year}_{month_name}_MNR.xlsx"


def _flatten_formulas_to_values(path: Path):
    """
    คืนค่า openpyxl Workbook (data_only=False) ที่ทุก formula cell ถูกเขียนทับด้วย cached value
    ของมันเองแล้ว (อ่านจาก pass แยกด้วย data_only=True) — ดู docstring หัวไฟล์สำหรับเหตุผล

    raises RuntimeError ถ้ามี formula cell ไหนไม่มี cached value ให้ fallback (ป้องกันการเขียนทับ
    ด้วย None โดยไม่รู้ตัว)
    """
    import openpyxl

    wb_vals = openpyxl.load_workbook(path, data_only=True)
    cache: dict[tuple[str, str], object] = {}
    for sheet_name in wb_vals.sheetnames:
        ws = wb_vals[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cache[(sheet_name, cell.coordinate)] = cell.value
    wb_vals.close()

    wb = openpyxl.load_workbook(path, data_only=False)
    missing = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    key = (sheet_name, cell.coordinate)
                    if key in cache:
                        cell.value = cache[key]
                    else:
                        missing.append(f"{sheet_name}!{cell.coordinate}")
    if missing:
        raise RuntimeError(
            f"พบ formula cell {len(missing)} จุดที่ไม่มี cached value ให้ fallback "
            f"(ตัวอย่าง: {missing[:10]}) -- หยุดก่อนเขียนทับเพื่อความปลอดภัย ไม่ flatten ไฟล์นี้"
        )
    return wb


def write_computed_days(
    target_dates: list[dt.date],
    sheet_source: Optional[str] = None,
    release_csv: Optional[Path] = None,
    dry_run: bool = False,
) -> list[dict]:
    """
    คำนวณ Inflow รายวันจากโทรมาตร (reservoir_daily_orchestration.compute_for_date()) แล้วเขียนผลลง
    ไฟล์ "บัญชีน้ำ" ทางการจริงของเดือนนั้นๆ (backup ไฟล์เดิมก่อนเสมอ ยกเว้น dry_run=True)

    release_csv=None (default) จะใช้ rdo.RELEASE_LOG_CSV -- ต้อง resolve ตอนเรียกจริง (ไม่ใช่ตอน
    import) เพราะไฟล์นี้กับ reservoir_daily_orchestration.py import กันเป็นวง (circular import) ถ้า
    ใช้ rdo.RELEASE_LOG_CSV เป็นค่า default ตรงๆ ตอน define ฟังก์ชันจะ error เพราะ rdo module ยัง
    init ไม่เสร็จตอนนั้น

    target_dates ต้องเป็นวันที่ในเดือนเดียวกันทั้งหมด (ไฟล์ .xlsx เป็นรายเดือน) — ถ้าคาบเกี่ยว 2 เดือน
    ให้เรียกฟังก์ชันนี้แยก 2 รอบ

    คืนค่า list ของ dict ผลลัพธ์แต่ละวัน (เหมือน reservoir_daily_orchestration.compute_for_date())
    เพิ่ม key "written_row" (แถวที่เขียนจริงในไฟล์) ถ้า dry_run=False
    """
    if not target_dates:
        return []
    if release_csv is None:
        release_csv = rdo.RELEASE_LOG_CSV
    months = {(d.year, d.month) for d in target_dates}
    if len(months) > 1:
        raise ValueError(f"target_dates คาบเกี่ยวหลายเดือน: {months} -- เรียกแยกทีละเดือน")

    path = official_file_path(target_dates[0])
    if not path.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ทางการ: {path}")

    results = []
    for target_date in sorted(target_dates):
        result = rdo.compute_for_date(target_date, sheet_source=sheet_source, release_csv=release_csv)
        # H (R runoff m3) ไม่ได้อยู่ใน compute_for_date() output ตรงๆ -- คำนวณซ้ำแบบเดียวกับที่
        # compute_daily_row() ทำภายใน (ใช้ level เดียวกัน, rain_mm เดียวกัน) เพื่อให้ตรงกับคอลัมน์ H
        curve_row = rwb._xlookup_floor(result["level_msl"], rwb._rating_curve())
        surface_area_m2 = curve_row[2]
        result["rain_runoff_m3"] = surface_area_m2 * ((result["rain_24h_mm"] or 0.0) / 1000.0)
        results.append(result)

    if dry_run:
        logger.info("dry_run=True -- ไม่เขียนไฟล์จริง (%s)", path)
        return results

    backup_path = path.with_name(
        path.name + f".bak_before_live_write_{dt.datetime.now():%Y%m%d_%H%M%S}"
    )
    shutil.copy2(path, backup_path)
    logger.info("backup ไฟล์เดิมไว้ที่ %s", backup_path)

    wb = _flatten_formulas_to_values(path)
    ws = wb[SHEET_NAME]

    for result in results:
        target_date = dt.date.fromisoformat(result["date"])
        row = FIRST_DATA_ROW + (target_date.day - 1)
        ws.cell(row=row, column=COL["day"], value=target_date.day)
        ws.cell(row=row, column=COL["level"], value=round(result["level_msl"], 3))
        ws.cell(row=row, column=COL["storage"], value=result["storage_m3"])
        ws.cell(row=row, column=COL["inflow"], value=result["inflow_m3"])
        ws.cell(row=row, column=COL["O"], value=result["release_o_m3"])
        ws.cell(row=row, column=COL["spill"], value=result["spill_m3"])
        ws.cell(row=row, column=COL["rain_mm"], value=result["rain_24h_mm"])
        ws.cell(row=row, column=COL["rain_m3"], value=result["rain_runoff_m3"])
        ws.cell(row=row, column=COL["evap"], value=result["evap_m3"])
        ws.cell(row=row, column=COL["infiltration"], value=result["infiltration_m3"])
        ws.cell(row=row, column=COL["delta_s"], value=result["delta_s_m3"])
        result["written_row"] = row
        logger.info(
            "เขียนแถว %d (%s): level=%.3f storage=%.1f inflow=%.1f data_complete=%s",
            row, result["date"], result["level_msl"], result["storage_m3"],
            result["inflow_m3"], result["data_complete"],
        )

    wb.save(path)
    logger.info("บันทึกไฟล์ %s เรียบร้อย (%d วัน)", path, len(results))
    return results


def main():
    import argparse
    import json

    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dates", type=str, required=True,
                         help="รายการวันที่ (YYYY-MM-DD) คั่นด้วยจุลภาค เช่น 2026-07-15,2026-07-16")
    parser.add_argument("--sheet-source", type=str, default=None)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    dates = [dt.date.fromisoformat(s.strip()) for s in args.dates.split(",")]
    results = write_computed_days(dates, sheet_source=args.sheet_source, dry_run=args.dry_run)
    print(json.dumps(results, indent=2, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
